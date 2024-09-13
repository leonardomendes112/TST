import pandas as pd
import time
import os
import numpy as np
from openpyxl import Workbook, load_workbook
import streamlit as st
from io import BytesIO
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import openpyxl.utils.cell
from openpyxl import load_workbook, Workbook
import os, time
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import pandas as pd
import re

st.set_page_config(page_title='🚍 Vehicle Distribution Report')
hide_streamlit_style = """
            <style>
            #MainMenu {visibility: hidden;}
            footer {visibility: hidden;}
            </style>
            """
st.markdown(hide_streamlit_style, unsafe_allow_html=True)

translation_dict = {
    "chapa": "duty id",
    "tipo de chapa": "duty type",
    "nome do motorista": "driver name",
    "id do motorista": "driver id",
    "placa": "board",
    "viatura": "vehicle",
    "início": "start",
    "término": "end",
    "de": "from",
    "para": "to",
    "linhas": "routes",
    "notas": "notes",
    "violações": "violations"
}

def remove_whitespaces(daily_report):

    ws = daily_report

    for row in ws.iter_rows(min_row=2, min_col=5, max_col=6):
        for cell in row:
            if cell.value != None:
                cell.value.replace(" ", "")

    return

def get_sheet_names(daily_report):
    wb = load_workbook(daily_report, data_only=True)

    return wb.sheetnames

def create_full_schedule_df(full_schedule):
    df = pd.read_excel(full_schedule)
    new_df = df[['Duty id', 'Vehicle Block Id', 'Start Time']].copy()
    new_df.dropna(subset=['Vehicle Block Id'], inplace=True)

    aux_dutyIdList = new_df['Duty id'].values.tolist()
    new_sequenceColumn = []
    for i in range(len(aux_dutyIdList)):
        new_sequenceColumn.append(i)

    new_df['Sequence'] = new_sequenceColumn

    new_df.sort_values(by=['Sequence'], inplace=True)

    duty_id = new_df['Duty id'].values.tolist()
    vehicle_block_id = new_df['Vehicle Block Id'].values.tolist()
    start_time = new_df['Start Time'].values.tolist()
    adj_start_time = [start_time[0]]

    for i in range(1, len(duty_id)):
        if duty_id[i] == duty_id[i-1] and vehicle_block_id[i] == vehicle_block_id[i-1]:
            if int(start_time[i].split(':')[0]) < int(adj_start_time[i-1].split(':')[0]):
                hour = int(start_time[i].split(':')[0])
                minute = start_time[i].split(':')[0]
                adj_hour = hour + 24
                adj_time = f"{adj_hour}:{minute}"
            else:
                adj_time = start_time[i]
        else:
            adj_time = start_time[i]
        adj_start_time.append(adj_time)

    new_df["adjusted_start_time"] = adj_start_time

    return new_df

def create_full_schedule_dict(full_schedule):

    df = create_full_schedule_df(full_schedule)
    df.columns = df.columns.str.lower().str.strip()
    df.sort_values(by=['vehicle block id','adjusted_start_time'], inplace=True)

    duty_id = df['duty id'].values.tolist()
    block_id = df['vehicle block id'].values.tolist()
    start_time = df['adjusted_start_time'].values.tolist()

    block_dict = {}

    for i in range(len(duty_id)):

        b_id = block_id[i]

        if b_id not in block_dict.keys():
            block_dict[b_id] = {}
            block_dict[b_id]['duty_id'] = []
            block_dict[b_id]['duty_id'].append(duty_id[i])
            block_dict[b_id]['start_time'] = []
            block_dict[b_id]['start_time'].append(start_time[i])
            block_dict[b_id]['counter'] = 1

            continue

        if b_id == block_id[i-1] and duty_id[i] != duty_id[i-1]:
            block_dict[b_id]['duty_id'].append(duty_id[i])
            block_dict[b_id]['start_time'].append(start_time[i])
            block_dict[b_id]['counter'] += 1

    return block_dict


def create_daily_report_dict(daily_report, sheet_name):

    df = pd.read_excel(daily_report, sheet_name=sheet_name, dtype=str, engine='openpyxl')
    df.columns = df.columns.str.lower().str.strip()
    df.rename(columns=translation_dict, inplace=True)
    df = df[df['duty type'] != 'custom']
    df = df.fillna("")

    daily_blocks_dict = {}

    for index, row in df.iterrows():
        blocksList = row['board']

        if blocksList not in [None, "missing", ""]:
            blocksList = blocksList.split(",")

            for _index, block_id in enumerate(blocksList):
                block_id = block_id.strip()
                match = re.match(r"([^\s(]+)", block_id)

                if match:
                    block_id = match.group(1)

                if block_id not in [None, "missing", ""]:
                    if block_id not in daily_blocks_dict.keys():
                        planned_vehicle_id = row['vehicle']
                        if planned_vehicle_id is not None:
                            planned_vehicle_id = planned_vehicle_id.split(",")
                            if len(planned_vehicle_id) > 1:
                                try:
                                    planned_vehicle = planned_vehicle_id[_index].replace(" ", "")
                                except IndexError:
                                    planned_vehicle = None
                            else:
                                planned_vehicle = planned_vehicle_id[0]
                        else:
                            planned_vehicle = None

                        daily_blocks_dict[block_id] = {
                            "planned_vehicle": planned_vehicle,
                            "duty_id": [],
                            "drivers_name": [],
                            "start_time": [],
                            "vehicle_id": []
                        }

                    duty_id = row['duty id']
                    start_time = row['start']
                    drivers_name = row['driver name']
                    try:
                        adjusted_drivers_name = f"{drivers_name.split(' ')[0].capitalize()} {drivers_name.split(' ')[-1].capitalize()}"
                    except:
                        adjusted_drivers_name = ""
                    vehicle_id = ""

                    daily_blocks_dict[block_id]["duty_id"].append(duty_id)
                    daily_blocks_dict[block_id]["start_time"].append(start_time)
                    daily_blocks_dict[block_id]["drivers_name"].append(adjusted_drivers_name)
                    daily_blocks_dict[block_id]["vehicle_id"].append(vehicle_id)

    return daily_blocks_dict



def create_table_lines_list(daily_blocks_dict, fs_blocks_dict):

    #These lists will store the ordered values so we can put them into a dataframe later on
    #The dataframe will be used for sorting purposes
    blockIdList = []
    dutyIdList = []
    driversNameList = []
    startTimeList = []

    #planned vehicle is always none (except for the planned Block instance)
    #iterate throught the daily report dict

    for block_id in daily_blocks_dict.keys():

        daily_DutyIdList = daily_blocks_dict[block_id]["duty_id"]
        daily_DriversNameList = daily_blocks_dict[block_id]["drivers_name"]
        daily_StartTimeList = daily_blocks_dict[block_id]["start_time"]

        #Iterate through the daily_DutyIdList

        for daily_index, dailyDutyId in enumerate(daily_DutyIdList):

            #Check if it is not a split duty

            if " (" not in dailyDutyId:

                #Iterate through the fullschedule dict for that block ID
                fs_DutyIdList = fs_blocks_dict[block_id]["duty_id"]

                for fs_index, fsDutyId in enumerate(fs_DutyIdList):

                    if dailyDutyId == fsDutyId:

                        fs_StartTime = fs_blocks_dict[block_id]["start_time"][fs_index]

                        blockIdList.append(block_id)
                        dutyIdList.append(dailyDutyId)
                        driversNameList.append(daily_DriversNameList[daily_index])
                        startTimeList.append(fs_StartTime)

            #If it is a case of split duty
            else:

                blockIdList.append(block_id)
                dutyIdList.append(dailyDutyId)
                driversNameList.append(daily_DriversNameList[daily_index])
                startTimeList.append(daily_StartTimeList[daily_index])

    df = pd.DataFrame(
        {"block_id" : blockIdList,
        "duty_id" : dutyIdList,
        "driver_name" : driversNameList,
        "start_time" : startTimeList
        })

    return df

#Create another dictionary called blockLinesDictionary to store the printable lines
def create_block_lines_dict(daily_blocks_dict, fs_blocks_dict):

    df = create_table_lines_list(daily_blocks_dict, fs_blocks_dict)

    df.sort_values(by=['block_id','start_time'], inplace=True)

    blockLinesDictionary = {}

    for index, row in df.iterrows():
        block_id = row["block_id"]

        if block_id not in blockLinesDictionary.keys():
            blockLinesDictionary[block_id] = {"count" : 0, "vehicle" : "", "line" : []}
            planned_vehicle = daily_blocks_dict[block_id]["planned_vehicle"]

            #Gets the planned vehicle ID from the daily report
            blockLinesDictionary[block_id]["vehicle"] = planned_vehicle

            #Appends the first columns into the final line
            blockLinesDictionary[block_id]["line"].extend([block_id, blockLinesDictionary[block_id]["vehicle"], ""])

        blockLinesDictionary[block_id]["count"] += 1

        duty_id = row["duty_id"]
        start_time = row["start_time"]
        drivers_name = row["driver_name"]
        start_time = row["start_time"]
        actual_vehicle = ""

        aux_line = [duty_id, start_time, drivers_name, actual_vehicle]

        blockLinesDictionary[block_id]["line"].extend(aux_line)

    return blockLinesDictionary

# Get the highest count of block occurrences and returns the max number
def get_max_count(blockLinesDictionary):
    countList = []

    for block_id in blockLinesDictionary.keys():
        countList.append(blockLinesDictionary[block_id]["count"])

    maxCount = max(countList)

    return maxCount

# Create output directory and returns its path
def create_output_directory(daily_report):

    timestamp = time.strftime("%Y%m%d%H%M%S")
    inputFilename = os.path.basename(daily_report).split(".")[0]

    newDirname = f"Reports_{inputFilename}_{timestamp}"

    inputFiledir = os.path.dirname(daily_report)
    newDirpath = os.path.join(inputFiledir, newDirname)

    os.mkdir(newDirpath)

    return newDirpath

# Create output file and returns its path
def create_output_file(newDirpath, sheet_name):
    timestamp = time.strftime("%Y_%m_%d-%H_%M_%S")
    fileName = f"Distribuicao_Viaturas_{sheet_name}_{timestamp}.xlsx"
    outputPath = os.path.join(newDirpath, fileName)

    wb = Workbook()
    #wb.save(outputPath)

    return outputPath

# Create header
def create_header(blockLinesDictionary):

    maxCount = get_max_count(blockLinesDictionary)

    baseHeader = ["Placa", "Viatura Programada", "Viatura"]

    recurrentHeader = ["Chapa", "Hora entrada", "Nome Motorista", "Viatura"]

    baseHeader.extend(maxCount * recurrentHeader)

    return baseHeader

# Populate header in output file
def print_header(wb, blockLinesDictionary):

    baseHeader = create_header(blockLinesDictionary)
    ws = wb.active
    ws.append(["DISTRIBUIÇÃO VIATURAS"])
    ws.append([""])
    ws.append([""])
    ws.append(baseHeader)

    return

# Populate lines in output file

def print_table_lines(wb, blockLinesDictionary):

    ws = wb.active

    for block_id in blockLinesDictionary.keys():
        ws.append(blockLinesDictionary[block_id]["line"])

    return

# Calculate table length
def get_table_length(wb):

    ws = wb.active

    tableLength = 0

    for row in ws.iter_rows(min_row=4, max_row=4):
        for _ in row:
            tableLength += 1

    return tableLength

# Calculate table height
def get_table_height(wb):

    ws = wb.active

    tableHeight = 0

    for row in ws.iter_rows(min_row=4, min_col=1, max_col=1):
        for _ in row:
            tableHeight += 1

    return tableHeight

# Style the output file

def style_output_report(wb):

    ws = wb.active

    titleFont = Font(name='Calibri',
                        size=11,
                        bold=True)

    titleAlignment = cellAlignment = Alignment(horizontal="center",
                        vertical="center")

    tableLength = get_table_length(wb)

    # Bold table title & merge cells
    tableTitle = ws.cell(row=1, column=1)
    tableTitle.font = titleFont
    tableTitle.alignment = titleAlignment

    # Merge title cells
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=tableLength)

    # Apply border to every cell in table
    build_borders(wb, cellAlignment, titleFont)

    #Adjust columns width
    format_columns_size(wb)

    return

def build_borders(wb, cellAlignment, titleFont):

    ws = wb.active

    tableLength = get_table_length(wb)

    tableHeight = get_table_height(wb)

    thinBorder = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
    #Top Left Header border or Chapa
    upperLeftborder = Border(left=Side(style='medium'),
                        top=Side(style='medium'))

    #Top Right Header border
    upperRightborder = Border(left=Side(style='thin'),
                        right=Side(style='medium'),
                        top=Side(style='medium'))

    #The very first top left cell below the header
    firstRowTopLeftborder = Border(left=Side(style='medium'),
                        top=Side(style='medium'))

    #The very last top right cell below the header
    firstRowTopRightborder = Border(left=Side(style='thin'),
                        right=Side(style='medium'),
                        top=Side(style='medium'))

    #Regular first row cell
    firstRowborder = Border(left=Side(style='thin'),
                        top=Side(style='medium'))

    #Top row Cell under Chapa
    TopCellUnderChapaborder = Border(left=Side(style='medium'),
                        top=Side(style='medium'))

    #Regular Cell under Chapa or Placa
    cellUnderChapaborder = Border(left=Side(style='medium'),
                        top=Side(style='thin'))

    #Regular cell
    regularCellborder =  Border(left=Side(style='thin'),
                        top=Side(style='thin'))

    #End of row cell
    endOfRowborder = Border(left=Side(style='thin'),
                        top=Side(style='thin'),
                        right=Side(style='medium'))

    #BottomLeftRowcell
    bottomLeftRowCellborder = Border(left=Side(style='medium'),
                        top=Side(style='thin'),
                        bottom=Side(style='medium'))

    #BottomRowcell
    bottomRowCellborder = Border(left=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='medium'))

    #BottomRightRowcell
    bottomRightRowCellborder = Border(left=Side(style='thin'),
                        right=Side(style='medium'),
                        top=Side(style='thin'),
                        bottom=Side(style='medium'))

    #BottomRowChapaCell
    bottomRowChapaCellborder = Border(left=Side(style='medium'),
                        top=Side(style='thin'),
                        bottom=Side(style='medium'))

    # Applying borders to header only
    for row in ws.iter_rows(min_row=4, max_row=4):
        for cell in row:
            cell.font = titleFont
            cell.alignment = cellAlignment
            if cell.value in ["Placa", "Chapa"]:
                cell.border = upperLeftborder
            elif cell.column == tableLength:
                cell.border = upperRightborder
            else:
                cell.border = firstRowborder

    # Applying borders to first line only
    for row in ws.iter_rows(min_row=5, max_row=5):
        for cell in row:
            cell.alignment = cellAlignment
            correspHeaderCell = ws.cell(row = 4, column = cell.column)
            if correspHeaderCell.value == "Placa":
                cell.border = firstRowTopLeftborder
            elif correspHeaderCell.value == "Chapa":
                cell.border = TopCellUnderChapaborder
            elif cell.column == tableLength:
                cell.border = firstRowTopRightborder
            else:
                cell.border = firstRowborder

    #Applying borders to middle lines only (all but the last)
    for row in ws.iter_rows(min_row=6, max_row = tableHeight-2+4 ):
        for cell in row:
            cell.alignment = cellAlignment
            correspHeaderCell = ws.cell(row = 4, column = cell.column)
            if correspHeaderCell.value in ["Placa", "Chapa"]:
                cell.border = cellUnderChapaborder
            elif cell.column == tableLength:
                cell.border =  endOfRowborder
            else:
                cell.border = regularCellborder

    #Applying border to the last row only
    for row in ws.iter_rows(min_row=tableHeight+4-1, max_row=tableHeight+4-1):
        for cell in row:
            cell.alignment = cellAlignment
            correspHeaderCell = ws.cell(row = 4, column = cell.column)
            if correspHeaderCell.value in ["Placa", "Chapa"]:
                cell.border = bottomRowChapaCellborder
            elif cell.column == tableLength:
                cell.border = bottomRightRowCellborder
            else:
                cell.border = bottomRowCellborder
    return

def format_columns_size(wb):

    ws = wb.active

    ws.column_dimensions['A'].width = "7"
    ws.column_dimensions['B'].width = "6"
    ws.column_dimensions['C'].width = "11"
    ws.column_dimensions['D'].width = "11"
    ws.column_dimensions['E'].width = "13"
    ws.column_dimensions['F'].width = "6"
    ws.column_dimensions['G'].width = "11"
    ws.column_dimensions['H'].width = "11"
    ws.column_dimensions['I'].width = "13"
    ws.column_dimensions['J'].width = "6"
    ws.column_dimensions['K'].width = "11"
    ws.column_dimensions['L'].width = "11"
    ws.column_dimensions['M'].width = "13"
    ws.column_dimensions['N'].width = "6"

    return

def build_vehicle_reports(full_schedule, daily_report):

    sheetNames = get_sheet_names(daily_report)

    fs_blocks_dict = create_full_schedule_dict(full_schedule)

    outputFilesList = []

    counter = 0

    timestamp = time.strftime("%Y%m%d%H%M%S")

    for sheet_name in sheetNames:

        _fileName = "{}_{}.xlsx".format(sheet_name , timestamp)
        outputPath = BytesIO()
        wb = Workbook()
        counter += 1
        daily_blocks_dict = create_daily_report_dict(daily_report, sheet_name)
        blockLinesDictionary = create_block_lines_dict(daily_blocks_dict, fs_blocks_dict)
        
        print_header(wb, blockLinesDictionary)
        print_table_lines(wb, blockLinesDictionary)
        style_output_report(wb)
        wb.save(outputPath)
        
        st.download_button(label= f'Download Report {sheet_name}', data=outputPath, file_name=f'Vehicle_Distribution_Report_'+f'{_fileName}')

    return

st.subheader('🚍 Vehicle Distribution Report')
full_schedule = st.file_uploader('Upload the Full Schedule', type= 'xlsx')
daily_report = st.file_uploader('Upload the Daily Report', type= 'xlsx')

if full_schedule and daily_report:

    build_vehicle_reports(full_schedule, daily_report)
