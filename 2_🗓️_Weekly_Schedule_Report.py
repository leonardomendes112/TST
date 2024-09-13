import calendar
import datetime
import os
import time
from io import BytesIO
import numpy as np
import openpyxl.utils.cell
import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Protection,
                             Side)

st.set_page_config(page_title='🗓️ Weekly Schedule Report')
hide_streamlit_style = """
            <style>
            #MainMenu {visibility: hidden;}
            footer {visibility: hidden;}
            </style>
            """
st.markdown(hide_streamlit_style, unsafe_allow_html=True)

def validate_daily_report(df):
    required_columns = ['DUTY ID', 'DUTY TYPE', 'DRIVER NAME', 'DRIVER ID', 'BOARD', 'VEHICLE', 'START', 'END', 'FROM', 'TO', 'ROUTES', 'NOTES', 'VIOLATIONS']

    uppercase_columns = set(required_columns)

    missing_columns = uppercase_columns.difference(set(df.columns.str.upper()))

    if missing_columns:
        error_msg = 'The following required columns are missing from the DataFrame: {}'.format(', '.join(missing_columns))
        raise ValueError(error_msg)

    actual_columns = list(df.columns.str.upper())

    if required_columns != actual_columns[:len(required_columns)]:
        error_msg = 'The order of the required columns is not correct. Required order is: {}, Actual order is: {}'.format(required_columns, actual_columns)
        raise ValueError(error_msg)


def translate_headers(df):
    df.rename(columns={
        'CHAPA': 'DUTY ID',
        'TIPO DE CHAPA': 'DUTY TYPE',
        'NOME DO MOTORISTA': 'DRIVER NAME',
        'ID DO MOTORISTA': 'DRIVER ID',
        'PLACA': 'BOARD',
        'VIATURA': 'VEHICLE',
        'INÍCIO': 'START',
        'TÉRMINO': 'END',
        'DE': 'FROM',
        'PARA': 'TO',
        'LINHAS': 'ROUTES',
        'NOTAS': 'NOTES',
        'VIOLAÇÕES': 'VIOLATIONS'}, inplace=True)

    return df

def get_weekday_name(day):
    day = day.lower()
    dict_weekday = {
        'monday' : 'segunda',
        'tuesday' : 'terca',
        'wednesday' : 'quarta',
        'thursday' : 'quinta',
        'friday' : 'sexta',
        'saturday' : 'sabado',
        'sunday' : 'domingo',
    }
    return dict_weekday[day]

def get_month_name_pt(month):
    dict_month = {
        '01' : 'janeiro',
        '02' : 'fevereiro',
        '03' : 'março',
        '04' : 'abril',
        '05' : 'maio',
        '06' : 'junho',
        '07' : 'julho',
        '08' : 'agosto',
        '09' : 'setembro',
        '10' : 'outubro',
        '11' : 'novembro',
        '12' : 'dezembro',
    }
    return dict_month[month]
    

def get_sheet_list(input_file):
    wb = load_workbook(input_file)
    sheetList = wb.sheetnames
    
    return sheetList

def drivers_count_dict(ws):
    drivers_count_dict = {}

    for row in ws.iter_rows(min_row = 2, min_col = 4, max_col = 4):
        for cell in row:
            driver_id = cell.value
            if driver_id != None:
                if driver_id not in drivers_count_dict.keys():
                    drivers_count_dict[driver_id] = 0
                drivers_count_dict[driver_id] += 1
    
    return drivers_count_dict

def get_output_path(inputFile,sheet_name):
    timestamp = time.strftime("%Y%m%d-%H%M%S")
    outputFile = "OPS_Tabela_de_Escala_" + sheet_name + '_' + timestamp + '.xlsx'
    
    return outputFile

def get_sheet_data(input_file, sheet_name):
    wb = load_workbook(input_file, data_only=True)
    ws = wb[sheet_name]

    drivers_count = drivers_count_dict(ws)

    df = pd.read_excel(input_file, sheet_name=sheet_name, dtype=str)

    df = df.rename(columns=str.upper)

    if df.columns.str.contains(r'Motorista', case=False, regex=True).any():
        translate_headers(df)

    validate_daily_report(df)
    df['DRIVER NAME'] = df['DRIVER NAME'].str.upper()

    unassigned_df = df.copy()

    df[['DRIVER ID']] = df[['DRIVER ID']].fillna(value='')
    df[['VEHICLE']] = df[['VEHICLE']].fillna(value='')
    df[['BOARD']] = df[['BOARD']].fillna(value='')
    df[['START']] = df[['START']].fillna(value='')
    df[['END']] = df[['END']].fillna(value='')
    df[['FROM']] = df[['FROM']].fillna(value='')

    df.sort_values(by=['DRIVER NAME', 'START'], inplace=True)

    data_list = []

    for index, row in df.iterrows():
        row_list = []

        driver_id = row['DRIVER ID']
        duty_type = row['DUTY TYPE']

        if driver_id != '':

            if drivers_count[driver_id] > 1 and duty_type == "day_off":
                continue
            
            row_list.append(row['DRIVER ID'])
            driver_name = row['DRIVER NAME'].split(' ')
            row_list.append(f"{driver_name[0]} {driver_name[-1]}")
            row_list.append(adjust_content(row['DUTY ID']))
            row_list.append(adjust_content(row['VEHICLE']))
            row_list.append(adjust_content(row['START']))
            row_list.append(adjust_content(row['END']))
            if row['DUTY TYPE'] == "day_off":
                row_list.append("Descanso (Folga)")
            else:
                row_list.append('')
            
            data_list.append(row_list)

    # Now appending the unassigned duties

    unassigned_df = unassigned_df[unassigned_df['DRIVER ID'].isnull()]

    for index, row in unassigned_df.iterrows():
        row_list = []
        row_list.append(row['DRIVER ID'])
        row_list.append(row['DRIVER NAME'])
        row_list.append(adjust_content(row['DUTY ID']))
        row_list.append(adjust_content(row['VEHICLE']))
        row_list.append(adjust_content(row['START']))
        row_list.append(adjust_content(row['END']))
        row_list.append('Chapa não escalada')
        
        data_list.append(row_list)

    #st.write(unassigned_df)
    
    return data_list

def print_sheet_data(data_list, output_file):
    wb = load_workbook(output_file, data_only=True)
    ws = wb.active
    
    for row in data_list:
        ws.append(row)
    
    wb.save(output_file)

    return

def print_header(output_file):
    HEADER = ['NÚMERO', 'NOME', 'CHAPA', 'VIAT', 'INÍCIO', 'FIM', 'OBSERVACOES']
    wb = load_workbook(output_file, data_only=True)
    ws = wb.active
    ws.append(HEADER)

    wb.save(output_file)

    return


def adjust_date_char(date_str):
    if len(str(date_str)) == 1:
        date_str = '0' + str(date_str)
    
    return date_str

def adjust_datetime(date_str):
    format_str = '%Y-%m-%d'
    datetime_obj = datetime.datetime.strptime(date_str, format_str)
    day = adjust_date_char(datetime_obj.day)
    month = adjust_date_char(datetime_obj.month)
    print(month)
    year = datetime_obj.year

    adjusted_date = f"{day} DE {get_month_name_pt(str(month)).upper()} DE {year} ({(get_weekday_name(calendar.day_name[datetime_obj.weekday()])).upper()})"

    return adjusted_date

def build_weekly_schedule_report(daily_report):

    sheet_list = get_sheet_list(daily_report)

    for sheet in sheet_list:

        outputFileName = get_output_path(daily_report, sheet)

        outputFile = BytesIO()
        wb = Workbook()
        wb.save(outputFile)

        ws = wb.active
        ws.title = f"{sheet}"

        wb.save(outputFile)

        print_header(outputFile)

        data_list = get_sheet_data(daily_report, sheet)

        print_sheet_data(data_list, outputFile)

        style_sheet(outputFile, sheet)

        outputFile.seek(0)
        wb.save(outputFile)

        st.download_button(label=f'Download Report {outputFileName}', data=outputFile.getvalue(), file_name=outputFileName, mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    return

def adjust_content(input_variable):
    content_dict = {
        "day_off" : "",
        "missing" : "",
        None: ""
    }

    if input_variable in content_dict.keys():
        adjusted_variable = content_dict[input_variable]
        return adjusted_variable
    
    return input_variable

def delete_redundant_dayoffs(ws):
    driver_count_dict = {}

    for row in ws.iter_rows(min_row = 2, min_col = 4, max_col = 4):
        for cell in row:
            driver_id = cell.value
            if driver_id != None:
                if driver_id not in driver_count_dict.keys():
                    driver_count_dict[driver_id] = 0
                driver_count_dict[driver_id] += 1

    for row in ws.iter_rows(min_row = 2, min_col = 4, max_col = 4):
        for cell in row:
            driver_id = cell.value
            duty_type = ws.cell(row = row[0].row, column = 2).value
            if driver_id != None:
                if driver_count_dict[driver_id] > 1 and duty_type == "day_off":
                    ws.delete_rows(row[0].row, 1)

    return

def style_sheet(output_file, sheet):

    wb = load_workbook(output_file)

    ws = wb.active

    ws.insert_rows(1,6)

    ws.cell(row = 1, column = 1).value = 'TST - TRANSPORTES SUL'
    ws.cell(row = 2, column = 1).value = 'GP24745'
    
    ws.cell(row = 3, column = 1).value = 'ESCALA DE SERVICO - AFECTACAO DE CHAPAS'
    ws.cell(row = 3, column = 1).font = Font(bold = True)
    ws.cell(row = 3, column = 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text = False)
    
    ws.cell(row = 4, column = 1).value = 'ordem alfabética'
    ws.cell(row = 4, column = 1).alignment = Alignment(horizontal='right', vertical='center', wrap_text = False)

    ws.cell(row = 5, column = 1).value = f'SERVICO A EFECTUAR NO DIA {adjust_datetime(sheet)}'
    ws.cell(row = 5, column = 1).alignment = Alignment(horizontal='left', vertical='center', wrap_text = False)
    ws.cell(row = 5, column = 1).font = Font(bold = True)
    
    ws.merge_cells('A3:G3')
    ws.merge_cells('A4:G4')
    ws.merge_cells('A5:G5')

    ws.column_dimensions['A'].width = "12.17"
    ws.column_dimensions['B'].width = "23.17"
    ws.column_dimensions['C'].width = "14.67"
    ws.column_dimensions['D'].width = "14.67"
    ws.column_dimensions['E'].width = "6.17"
    ws.column_dimensions['F'].width = "6.17"
    ws.column_dimensions['G'].width = "52.17"

    dash_border = Border(left=Side(style='dashed'), 
                     right=Side(style='dashed'), 
                     top=Side(style='dashed'), 
                     bottom=Side(style='dashed'))

    for row in ws.iter_rows(min_row = 7, max_col = 7):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text = False)
            cell.border = dash_border

    for row in ws.iter_rows(min_row=8, max_col=7):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text = False)

    wb.save(output_file)

    return

st.subheader('🗓️ Weekly Schedule Report')
daily_report = st.file_uploader('Upload the Daily Report', type= 'xlsx')

if daily_report:
    build_weekly_schedule_report(daily_report)