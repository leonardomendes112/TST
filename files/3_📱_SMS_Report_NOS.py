import csv
import datetime
import numpy as np
import os
import pandas as pd
import tempfile
import time

from openpyxl import Workbook, load_workbook

import streamlit as st

from io import BytesIO
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl import Workbook

st.set_page_config(page_title='📱 SMS Report')
hide_streamlit_style = """
            <style>
            #MainMenu {visibility: hidden;}
            footer {visibility: hidden;}
            </style>
            """
st.markdown(hide_streamlit_style, unsafe_allow_html=True)

def get_telephone_file(input_file):
    drivers_phones = {}

    driversDf = pd.read_csv(input_file, dtype=str)
    driversDf["Mobile Number"].fillna(driversDf["Home Number"], inplace=True)

    driversDf.set_index("ID", inplace=True)
    drivers_phones = driversDf.to_dict("index")
    
    return drivers_phones

def get_output_path(inputFile,sheet_name):
    parentDir = os.path.dirname(inputFile)
    reportName = os.path.basename(inputFile).split('.')[0]
    timestamp = time.strftime("%Y%m%d-%H%M%S")
    outputFile = "OPS_SMS_Report_NOS_"+ reportName + '_' + sheet_name + '_' + timestamp + '.xlsx'
    outputPath = os.path.join(parentDir, outputFile)
    
    return outputPath

def get_sheet_list(input_file):
    wb = load_workbook(input_file)
    sheetList = wb.sheetnames
    
    return sheetList

def translate_headers(df):
    df.rename(columns={
        'CHAPA': 'DUTY ID',
        'TIPO DE CHAPA': 'DUTY TYPE',
        'NOME DO MOTORISTA': 'DRIVER NAME',
        'ID DO MOTORISTA': 'DRIVER ID',
        'PLACA': 'BOARD',
        'VIATURA': 'VEHICLE',
        'INÍCIO': 'START',
        'TÉRMINO': 'END',
        'DE': 'FROM',
        'PARA': 'TO',
        'LINHAS': 'ROUTES',
        'NOTAS': 'NOTES',
        'VIOLAÇÕES': 'VIOLATIONS'}, inplace=True)

    return df


def get_sheet_data(input_file, sheet_name, drivers_phones):
    wb = load_workbook(input_file, data_only=True)
    ws = wb[sheet_name]

    drivers_count = drivers_count_dict(ws)

    daily_dict = {}

    df = pd.read_excel(input_file, sheet_name=sheet_name, dtype=str)

    df = df.rename(columns=str.upper)
    
    if "CHAPA" in df.columns:
        translate_headers(df)

    df[['DRIVER ID']] = df[['DRIVER ID']].fillna(value='xxx')
    df[['VEHICLE']] = df[['VEHICLE']].fillna(value='')
    df[['BOARD']] = df[['BOARD']].fillna(value='')
    df[['START']] = df[['START']].fillna(value='')
    df[['END']] = df[['END']].fillna(value='')
    df[['FROM']] = df[['FROM']].fillna(value='')

    df.sort_values(by=['START'], inplace=True)

    for index, row in df.iterrows():
        driver_id = row['DRIVER ID']
        
        duty_type = row['DUTY TYPE']
        duty_id = row['DUTY ID']


        if driver_id not in [None, '0', 0, 'xxx']:
        
            if drivers_count[driver_id] > 1 and duty_type == "day_off":
        
                continue

            if driver_id not in daily_dict.keys():
                daily_dict[driver_id] = {}
                if driver_id not in drivers_phones.keys():
                    daily_dict[driver_id]['TELEPHONE'] = ''
                else:
                    daily_dict[driver_id]['TELEPHONE'] = drivers_phones[driver_id]["Mobile Number"]

                daily_dict[driver_id]['SERVICE DATE'] = adjust_datetime(sheet_name)
                daily_dict[driver_id]['BLOCK'] = [adjust_content(row['BOARD'])]
                daily_dict[driver_id]['DUTY'] = [adjust_content(row['DUTY ID'])]
                daily_dict[driver_id]['VEHICLE'] = [str(adjust_content(row['VEHICLE']))]
                if row['START'] == None:
                    daily_dict[driver_id]['START AND END HOUR'] = [f" "]
                else:
                    daily_dict[driver_id]['START AND END HOUR'] = [f"{adjust_content(row['START'])}-{row['END']}"]
                daily_dict[driver_id]['DEPOT'] = [adjust_content(row['FROM'])]
                daily_dict[driver_id]['NOTES'] = []
                daily_dict[driver_id]['COMPANY'] = 'TST'
            else:
                daily_dict[driver_id]['NOTES'].append(f"Placa:{adjust_content(row['BOARD'])}  Chapa:{adjust_content(row['DUTY ID'])}  Viat:{adjust_content(row['VEHICLE'])}  Hor:{adjust_content(row['START'])} {adjust_content(row['END'])} Em:{adjust_content(row['FROM'])}")

            data_list = []

            new_data_list = [] #This one is used for the NOS version

            for driver_id in daily_dict.keys():
                row_list = []
                row_list.append(daily_dict[driver_id]['TELEPHONE'])
                row_list.append(f"Em:{daily_dict[driver_id]['SERVICE DATE']} ")
                row_list.append("Placa:" + ", ".join(str(item) for item in daily_dict[driver_id]['BLOCK']))
                row_list.append("Chapa:" + ", ".join(str(item) for item in daily_dict[driver_id]['DUTY']))
                row_list.append("Viat:" + ", ".join(str(item) for item in daily_dict[driver_id]['VEHICLE']))
                row_list.append("Hor:" + "*".join(str(item) for item in daily_dict[driver_id]['START AND END HOUR']))
                row_list.append("Em:" + ", ".join(str(item) for item in daily_dict[driver_id]['DEPOT']))
                row_list.append(" / ".join(str(item) for item in daily_dict[driver_id]['NOTES']))
                row_list.append(daily_dict[driver_id]['COMPANY'])

                new_telefone = daily_dict[driver_id]['TELEPHONE']
                new_service_date = f"Em:{daily_dict[driver_id]['SERVICE DATE']} "
                new_block = "Placa:" + ", ".join(str(item) for item in daily_dict[driver_id]['BLOCK'])
                new_duty = "Chapa:" + ", ".join(str(item) for item in daily_dict[driver_id]['DUTY'])
                new_vehicle = "Viat:" + ", ".join(str(item) for item in daily_dict[driver_id]['VEHICLE'])
                new_start_end_hour = "Hor:" + "*".join(str(item) for item in daily_dict[driver_id]['START AND END HOUR'])
                new_depot = "Em:" + ", ".join(str(item) for item in daily_dict[driver_id]['DEPOT'])
                new_notes = " / ".join(str(item) for item in daily_dict[driver_id]['NOTES'])
                new_company = daily_dict[driver_id]['COMPANY']

                
                data_list.append(row_list)

                mensagem = "{} {} {} {} {} {} {} {}".format(new_service_date, new_block, new_duty, new_vehicle, new_start_end_hour, new_depot, new_notes, new_company)
                tipo = ""
                destinatarios = new_telefone
                relatorio = "F"
                validade = "8"
                originador = "TST"
                enviarapartirde = "01-01-2022 00:00:00"

                new_data_list.append([mensagem, tipo, destinatarios, relatorio, validade, originador, enviarapartirde])
    
    return new_data_list

def print_sheet_data(data_list, df):
    
    for row in data_list:
        df.loc[len(df)] = row

    return df

#Prints the Header into the output file
def print_header(df):
    HEADER = ['Mensagem', 'Tipo', 'Destinatarios', 'Relatorio', 'Validade', 'Originador', 'Enviarapartirde'] #This is the new header for NOS

    df = pd.DataFrame(columns = HEADER)

    return df

#Returns a dictionary with the count for each DRIVER ID in a single sheet
def count_driverid(input_file, sheet_name):
    wb = load_workbook(input_file, data_only=True)
    ws = wb[sheet_name]

    drivers_count = {}
    for row in ws.iter_cols(min_row=2, min_col=4, max_col=4):
        for cell in row:
            if cell.value != None:
                driver_id = cell.value
                if driver_id not in drivers_count.keys():
                    drivers_count[driver_id] = 0
                drivers_count[driver_id] += 1

    for key in drivers_count.keys():
        print (drivers_count[key])
    return drivers_count

def adjust_date_char(date_str):
    if len(str(date_str)) == 1:
        date_str = '0' + str(date_str)
    
    return date_str

def adjust_datetime(date_str):
    format_str = '%Y-%m-%d'
    datetime_obj = datetime.datetime.strptime(date_str, format_str)
    day = adjust_date_char(datetime_obj.day)
    month = adjust_date_char(datetime_obj.month)
    year = datetime_obj.year

    adjusted_date = f"{day}/{month}/{year}"

    return adjusted_date

def build_sms_report(daily_report, drivers_csv):

    drivers_phones = get_telephone_file(drivers_csv)

    sheet_list = get_sheet_list(daily_report)


    for sheet in sheet_list:

        base_name = daily_report.name.split(".")[0]

        df = pd.DataFrame()

        df = print_header(df)

        data_list = get_sheet_data(daily_report, sheet, drivers_phones)

        df = print_sheet_data(data_list, df)

        downloadfile = BytesIO()

        df.to_excel(downloadfile, index=False)  

        st.download_button(label= f'Download Report {sheet}', data=downloadfile, file_name=f'SMS_Report_NOS_' + f'{sheet}.xlsx')

    return

def adjust_content(input_variable):
    content_dict = {
        "day_off" : "Folga",
        "missing" : "",
        None: ""
    }

    if input_variable in content_dict.keys():
        adjusted_variable = content_dict[input_variable]
        return adjusted_variable
    
    return input_variable

def drivers_count_dict(ws):
    drivers_count_dict = {}

    for row in ws.iter_rows(min_row = 2, min_col = 4, max_col = 4):
        for cell in row:
            driver_id = cell.value
            if driver_id != None:
                if driver_id not in drivers_count_dict.keys():
                    drivers_count_dict[driver_id] = 0
                drivers_count_dict[driver_id] += 1
    
    return drivers_count_dict
    
st.subheader('📱 SMS Report')
daily_report = st.file_uploader('Upload the Daily Report', type= 'xlsx')
driver_csv = st.file_uploader('Upload the Driver CSV', type= 'csv')

if daily_report and driver_csv:
    
    build_sms_report(daily_report, driver_csv)
